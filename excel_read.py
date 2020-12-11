import openpyxl

def readExcel():
    path = "py-read.xlsx"

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row=1, column=1)

    # Print total rows and columns in the sheet
    m_rows = sheet_obj.max_row
    m_columns = sheet_obj.max_column

    print(f"There are total : {m_rows} rows in the current sheet {sheet_obj}")
    print(f"There are total : {m_columns} columns in the current sheet {sheet_obj}")

    #Fetching name of the headings
    for i in range(1, m_columns+1):
      cell_obj = sheet_obj.cell(row = 1, column = i)
      print(cell_obj.value, end = " ")

    for i in range(1, m_columns+1):
      print(f"\nPrinting column {i}\n")
      for j in range(1, m_rows+1):
        cell_obj = sheet_obj.cell(row=j, column=i)
        print(cell_obj.value)